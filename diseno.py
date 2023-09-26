import descarga
import multiplex_probes
import os
import argparse
import pandas as pd
import primer3
import subprocess
import queue
from Bio.SeqUtils import GC, MeltingTemp
from Bio.Seq import Seq
from Bio import SearchIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


def check_probe(seq, iscentral, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple):
    """
    Función que recibe una sonda (secuencia) y las restricciones para verificar que la
    sonda cumpla los parámetros establecidos.
    Retorna un valor booleano que determina si cumple las restricciones o no.
    :param seq: La secuencia que se quiere verificar.
    :param minlen: El largo mínimo de la sonda.
    :param maxlen: El largo máximo de la sonda.
    :param tmmin: Temperatura de melting mínima.
    :param tmmax: Temperatura de melting máxima.
    :param gcmin: Porcentaje de GC mínimo.
    :param gcmax: Porcentaje de GC máximo.
    :param dgmin_homodim: Delta G mínimo permitido para validación de homodimerización.
    :param dgmin_hairpin: Delta G mínimo permitido para validación de hairpin u horquilla.
    :param
    :return: Booleano que determina su factibilidad
    """

    #Chequear largo
    largo = len(seq)
    if largo < minlen or largo > maxlen:
        #print('Largo no valido')
        return False
    
    #Chequear Tm
    tm = MeltingTemp.Tm_NN(seq)
    #tm = primer3.calc_tm(str(seq))
    if tm < tmmin or tm > tmmax:
        #print('Tm fuera de rango')
        return False
    
    #Chequear %GC
    if GC(seq) < gcmin or GC(seq) > gcmax:
        #print('GC fuera de rango')
        return False
    
    #Chequear hetero - homodimeros - horquilla
    #dg_hairpin = primer3.calcHairpin(str(seq)).dg
    if largo <= 60:
        if primer3.calcHairpin(str(seq)).dg < dgmin_hairpin or primer3.calcHomodimer(str(seq)).dg < dgmin_homodim:
            #print('Delta G')
            return False
    else:
        if primer3.calcHairpin(str(seq[:60])).dg < dgmin_hairpin or primer3.calcHomodimer(str(seq[:60])).dg < dgmin_homodim:
            #print('Delta G')
            return False

    
    # Chequear homopolimeros de 1 base
    bases = ['A', 'C', 'G', 'T']
    for b in bases:
        if b * (maxhomopol_simple + 1) in seq:
            #print('Homopolimero de 1 nucleotido')
            return False
        
    # Chequear homopolimeros de 2 bases
    checked = []
    for i in range(len(seq)-2):
        subseq = str(seq[i:i+1])
        if subseq not in checked:
            if subseq * (maxhomopol_double + 1) in seq:
                #print('Homopolimero de 2 nucleotidos')
                return False
            checked.append(subseq)

    # Chequear homopolimeros de 3 bases
    checked = []
    for i in range(len(seq)-3):
        subseq = str(seq[i:i+2])
        if subseq not in checked:
            if subseq * (maxhomopol_triple + 1) in seq:
                #print('Homopolimero de 3 nucleotidos')
                return False
            checked.append(subseq)
    
    #Chequear especificidad (BLAST LOCAL)
    #if not verify_specificity(seq, iscentral):
    #    print('Especificidad')
    #    return False
    
    #print('Sonda Aceptada: ', str(seq))
    return True


#Generar par de sondas para una ubicación y un site con ciertos parametros
def get_probes_from_pos(empalme, record, site, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, mindist, maxdist, minoverlap, maxoverlap, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple):
    """
    Esta función genera un par de sondas a partir de la secuencia de referencia y dos posiciones que representan el punto de empalme entre dos exones.
    :param empalme: Tupla con posiciones que representan al punto de empalme o splicing.
    :param record: Secuencia de referencia.
    :param site: Entero que representa si las sondas son donor, acceptor o central.
    :param minlen: El largo mínimo de la sonda.
    :param maxlen: El largo máximo de la sonda.
    :param tmmin: Temperatura de melting mínima.
    :param tmmax: Temperatura de melting máxima.
    :param gcmin: Porcentaje de GC mínimo.
    :param gcmin: Porcentaje de GC máximo.
    :param mindist: Distancia mínima al borde del exón.
    :param maxdist: Distancia máxima al borde del exón.
    :param minoverlap: Sobrelape mínimo entre sondas.
    :param maxoverlap: Sobrelape máximo entre sondas.
    :param dgmin_homodim: Delta G mínimo permitido para validación de homodimerización.
    :param dgmin_hairpin: Delta G mínimo permitido para validación de hairpin u horquilla.
    :
    :return: Lista con dos sondas, cada sonda es una tupla con la secuencia, la posición de inicio, la posición final y la distancia al borde del exón si aplica.
    """
    b = False
    if site == 0:
        i = minlen
        while i <= maxlen and not b:
            pos_inicio_int = empalme[0]-i//2
            pos_fin_int = empalme[1]+i//2
            probe_int = (record.seq[pos_inicio_int:empalme[0]] + record.seq[empalme[1]:pos_fin_int]).reverse_complement()
            b = check_probe(probe_int, True, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
            i += 1
        if not b:
            return [(Seq('AAA'), -1, -1, -1), (Seq('AAA'), -1, -1, -1)]
        else:
            b = False
            i +=1
            while i <= maxlen and not b:
                pos_inicio_ext = empalme[0]-i//2
                pos_fin_ext = empalme[1]+i//2
                probe_ext = (record.seq[pos_inicio_ext:empalme[0]] + record.seq[empalme[1]:pos_fin_ext]).reverse_complement()
                b = check_probe(probe_ext, True, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                i += 1
            if b:
                return [(probe_int, pos_inicio_int, pos_fin_int, -1), (probe_ext, pos_inicio_ext, pos_fin_ext, -1)]
            else:
                return [(probe_int, pos_inicio_int, pos_fin_int, -1), (Seq('AAA'), -1, -1, -1)]


    elif site == -1:
        dist_to_border = mindist
        while dist_to_border <= maxdist and not b:
            i = minlen
            while i <= maxlen and not b:
                pos_inicio_int = empalme[0]-i-dist_to_border
                pos_fin_int = empalme[0]-dist_to_border
                probe_int = record.seq[pos_inicio_int:pos_fin_int].reverse_complement()
                b = check_probe(probe_int, False, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                i += 1
            dist_to_border += 1
        if not b:
            return [(Seq('AAA'), -1, -1, -1), (Seq('AAA'), -1, -1, -1)]    
        else:
            b = False
            dist_int_probe = dist_to_border
            dist_to_border += i * (100-maxoverlap) / 100
            while dist_to_border <= dist_int_probe + i * (100-minoverlap) / 100 and not b:
                j = minlen
                while j <= maxlen and not b:
                    pos_inicio_ext = int(empalme[0]-j-dist_to_border)
                    pos_fin_ext = int(empalme[0]-dist_to_border)
                    probe_ext = record.seq[pos_inicio_ext:pos_fin_ext].reverse_complement()
                    b = check_probe(probe_ext, False, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                    j += 1
                dist_to_border += 1
            if b:
                return [(probe_int, pos_inicio_int, pos_fin_int, dist_int_probe), (probe_ext, pos_inicio_ext, pos_fin_ext, dist_to_border)]
            else:
                return [(probe_int, pos_inicio_int, pos_fin_int, dist_int_probe), (Seq('AAA'), -1, -1, -1)]
            

    elif site == 1:
        dist_to_border = mindist
        while dist_to_border <= maxdist and not b:
            i = minlen
            while i <= maxlen and not b:
                pos_inicio_int = empalme[1]+dist_to_border
                pos_fin_int = empalme[1]+i+dist_to_border
                probe_int = record.seq[pos_inicio_int:pos_fin_int].reverse_complement()
                b = check_probe(probe_int, False, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                i += 1
            dist_to_border += 1
        
        if not b:
            return [(Seq('AAA'), -1, -1, -1), (Seq('AAA'), -1, -1, -1)]    
        else:
            b = False
            dist_int_probe = dist_to_border
            dist_to_border += i * (100-maxoverlap) / 100
            while dist_to_border <= dist_int_probe + i * (100-minoverlap) / 100 and not b:
                j = minlen
                while j <= maxlen and not b:
                    pos_inicio_ext = int(empalme[1]+dist_to_border)
                    pos_fin_ext = int(empalme[1]+j+dist_to_border)
                    probe_ext = record.seq[pos_inicio_ext:pos_fin_ext].reverse_complement()
                    b = check_probe(probe_ext, False, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                    j += 1
                dist_to_border += 1
            if b:
                return [(probe_int, pos_inicio_int, pos_fin_int, dist_int_probe), (probe_ext, pos_inicio_ext, pos_fin_ext, dist_to_border)]
            else:
                return [(probe_int, pos_inicio_int, pos_fin_int, dist_int_probe), (Seq('AAA'), -1, -1, -1)]

def verify_specificity(seq, iscentral):
    """
    Función que verifica la especificidad o unicidad de una secuencia usando la herramienta de alineamiento BLAST de manera local.
    Es necesario tener instalado Blast+ y las bases de datos en la carpeta 'refseq'.
    :param seq: Secuencia de nuecleótidos que se quiere validar.
    :param iscentral: Booleano que indica si se trata de una sonda central. Se utilizan diferentes criterios para chequear la especificidad de una sonda central.
    :return: Bolleano que indica si la secuencia se considera específica.
    """
    db1_name = "refseq/db1_genoma"
    db2_name = "refseq/db2_transcriptoma"
    query = "refseq/query.txt"
    outxml = 'refseq/resultados.xml'

    with open(query, "w") as archivo:
        archivo.write(">1\n")
        archivo.write(str(seq))

    blastn_cmd = ['blastn',
                '-task', 'blastn-short',
                '-query', query,
                '-db', db1_name,
                '-out', outxml,
                '-outfmt', '5',
                '-word_size', '11',
                '-dust', 'no']

    subprocess.run(blastn_cmd, shell=True)
    blast_qresult = SearchIO.read("refseq/resultados.xml", "blast-xml")

    if iscentral:
        if blast_qresult[0].hsps[2].evalue / blast_qresult[0].hsps[1].evalue < 100:
            return False
        if blast_qresult[1].hsps[0].evalue / blast_qresult[0].hsps[1].evalue < 100:
            return False
        
    else:
        if blast_qresult[0].hsps[1].evalue / blast_qresult[0].hsps[0].evalue < 100:
            return False
        if blast_qresult[1].hsps[0].evalue / blast_qresult[0].hsps[0].evalue < 100:
            return False
    return True

#Obtener zonas de empalme
def get_splicing_pairs(locations):
    """
    Función que obtiene las zonas de empalme a partir de una serie de regiones que representan una transcripción.
    Las zonas de empalme se representan por dos posiciones: el fin de un exón y el inicio del siguiente.
    :param locations: Subregiones que representan la secuencia codificante o secuencia de exones
    :return: Lista con pares de posiciones en una tupla: el fin de un exón y el inicio del siguiente.
    """
    positions = []
    for subloc in locations.parts:
        positions.append((subloc.start, subloc.end))

    empalmes = []
    for i in range(len(positions)-1):
        empalmes.append((positions[i][1].position, positions[i+1][0].position))
    return empalmes

def get_splicings(seqrecord, transcripts):
    empalmes_array = []
    for f in seqrecord.features:
        if f.type == 'CDS':
            if f.qualifiers.get('product')[0] in transcripts:
                loc = f.location
                pares_empalme = get_splicing_pairs(loc)
                for par in pares_empalme:
                    if par not in empalmes_array:
                        empalmes_array.append(par)
    return empalmes_array

def probe_designer(record, transcripts, progreso_queue, minlen=60, maxlen=120, tmmin=65, tmmax=80, gcmin=30, gcmax=70, mindist=0, maxdist=200, minoverlap=25, maxoverlap=50, dgmin_homodim=-10000, dgmin_hairpin=-10000, maxhomopol_simple=6, maxhomopol_double=5, maxhomopol_triple=4, multiplex=True, mindg=-13627, maxdt=5):
    """
    Función principal diseñadora de sondas. De la secuancia anotada se obtiene una serie de sondas como subsecuencias de ésta, que cumplen varias restricciones ajustadas por los parámetros.
    :param record: Secuencia anotada de referencia para la creación de las sondas.
    :param transcripts: Transcripciones en las cuales se quiere diseñar.
    :param minlen: El largo mínimo de la sonda.
    :param maxlen: El largo máximo de la sonda.
    :param tmmin: Temperatura de melting mínima.
    :param tmmax: Temperatura de melting máxima.
    :param gcmin: Porcentaje de GC mínimo.
    :param gcmin: Porcentaje de GC máximo.
    :param mindist: Distancia mínima al borde del exón.
    :param maxdist: Distancia máxima al borde del exón.
    :param minoverlap: Sobrelape mínimo entre sondas.
    :param maxoverlap: Sobrelape máximo entre sondas.
    :param dgmin_homodim: Delta G mínimo permitido para validación de homodimerización.
    :param dgmin_hairpin: Delta G mínimo permitido para validación de hairpin u horquilla.
    :param maxhomopol_simple:
    :return: Lista con pares de posiciones en una tupla: el fin de un exón y el inicio del siguiente.
    """
    locations = []
    dict_cds = {
        'gen':[],
        'info':[],
        'empalmes':[]
    }

    for f in record.features:
        if f.type == 'CDS':
            loc = f.location
            info = f.qualifiers.get('product', ['Transctipto no identificado'])[0]
            if str(loc) not in locations and info in transcripts:
                locations.append(str(loc))
                dict_cds['gen'].append(f.qualifiers.get('gene', ['Gen no identificado'])[0])
                dict_cds['info'].append(info)
                dict_cds['empalmes'].append(get_splicing_pairs(loc))

    df_cds = pd.DataFrame(dict_cds)

    dict_probes = {
        'nombre':[],
        'gen':[],
        'transcrito':[],
        'sonda':[],
        'pos inicio':[],
        'pos fin':[],
        'empalme':[],
        'lado':[],
        'tm':[],
        'gc':[],
        'largo':[],
        'dist':[]
        #'deltag hairpin':[],
        #'deltag homodim':[]
    }

    splicing_pairs = get_splicings(record, transcripts)
    dict_generated_probes = {}
    num_splicings = len(splicing_pairs)

    progreso_actual = 0

    for index, row in df_cds.iterrows():
        for empalme in row['empalmes']:
            #Verificar si aún no se diseñan las sondas para ese empalme
            if str(empalme) not in dict_generated_probes:
                donor_probes = get_probes_from_pos(empalme, record, -1, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, mindist, maxdist, minoverlap, maxoverlap, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)

                #Sonda donor interior
                dict_probes['nombre'].append('DI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])               
                dict_probes['sonda'].append(str(donor_probes[0][0]))
                dict_probes['pos inicio'].append(donor_probes[0][1])
                dict_probes['pos fin'].append(donor_probes[0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('donor interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(donor_probes[0][0]))
                dict_probes['gc'].append(int(GC(donor_probes[0][0])))
                dict_probes['largo'].append(len(donor_probes[0][0]))
                dict_probes['dist'].append(int(donor_probes[0][3]))

                #Sonda donor exterior
                dict_probes['nombre'].append('DE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])                
                dict_probes['sonda'].append(str(donor_probes[1][0]))
                dict_probes['pos inicio'].append(donor_probes[1][1])
                dict_probes['pos fin'].append(donor_probes[1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('donor exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(donor_probes[1][0]))
                dict_probes['gc'].append(int(GC(donor_probes[1][0])))
                dict_probes['largo'].append(len(donor_probes[1][0]))
                dict_probes['dist'].append(int(donor_probes[1][3]))

                acceptor_probes = get_probes_from_pos(empalme, record, 1, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, mindist, maxdist, minoverlap, maxoverlap, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)

                #Sonda acceptor interior
                dict_probes['nombre'].append('AI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])                
                dict_probes['sonda'].append(str(acceptor_probes[0][0]))
                dict_probes['pos inicio'].append(acceptor_probes[0][1])
                dict_probes['pos fin'].append(acceptor_probes[0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('acceptor interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(acceptor_probes[0][0]))
                dict_probes['gc'].append(int(GC(acceptor_probes[0][0])))
                dict_probes['largo'].append(len(acceptor_probes[0][0]))
                dict_probes['dist'].append(int(acceptor_probes[0][3]))


                #Sonda acceptor exterior
                dict_probes['nombre'].append('AE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(acceptor_probes[1][0]))
                dict_probes['pos inicio'].append(acceptor_probes[1][1])
                dict_probes['pos fin'].append(acceptor_probes[1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('acceptor exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(acceptor_probes[1][0]))
                dict_probes['gc'].append(int(GC(acceptor_probes[1][0])))
                dict_probes['largo'].append(len(acceptor_probes[1][0]))
                dict_probes['dist'].append(int(acceptor_probes[1][3]))

                central_probes = get_probes_from_pos(empalme, record, 0, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, mindist, maxdist, minoverlap, maxoverlap, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)

                #Sonda central interior
                dict_probes['nombre'].append('CI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])    
                dict_probes['sonda'].append(str(central_probes[0][0]))
                dict_probes['pos inicio'].append(central_probes[0][1])
                dict_probes['pos fin'].append(central_probes[0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('central interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(central_probes[0][0]))
                dict_probes['gc'].append(int(GC(central_probes[0][0])))
                dict_probes['largo'].append(len(central_probes[0][0]))
                dict_probes['dist'].append(int(central_probes[0][3]))

                #Sonda central exterior
                dict_probes['nombre'].append('CE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])                
                dict_probes['sonda'].append(str(central_probes[1][0]))
                dict_probes['pos inicio'].append(central_probes[1][1])
                dict_probes['pos fin'].append(central_probes[1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('central exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(central_probes[1][0]))
                dict_probes['gc'].append(int(GC(central_probes[1][0])))
                dict_probes['largo'].append(len(central_probes[1][0]))
                dict_probes['dist'].append(int(central_probes[1][3]))

                dict_generated_probes[str(empalme)] = [donor_probes, acceptor_probes, central_probes]

                progreso_actual += int(100 / num_splicings)
                progreso_queue.put(progreso_actual)


            #Si el punto de empalme ya fue diseñado se extrae del diccionario
            else:
                #Sonda donor interior
                dict_probes['nombre'].append('DI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])                
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][0][0][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][0][0][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][0][0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('donor interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][0][0][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][0][0][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][0][0][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][0][0][3]))

                #Sonda donor exterior
                dict_probes['nombre'].append('DE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])               
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][0][1][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][0][1][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][0][1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('donor exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][0][1][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][0][1][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][0][1][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][0][1][3]))


                #Sonda acceptor interior
                dict_probes['nombre'].append('AI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][1][0][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][1][0][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][1][0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('acceptor interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][1][0][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][1][0][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][1][0][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][1][0][3]))

                #Sonda acceptor exterior
                dict_probes['nombre'].append('AE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][1][1][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][1][1][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][1][1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('acceptor exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][1][1][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][1][1][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][1][1][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][1][1][3]))


                #Sonda central interior
                dict_probes['nombre'].append('CI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][2][0][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][2][0][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][2][0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('central interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][2][0][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][2][0][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][2][0][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][2][0][3]))


                #Sonda central exterior
                dict_probes['nombre'].append('CE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][2][1][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][2][1][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][2][1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('central exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][2][1][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][2][1][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][2][1][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][2][1][3]))


    df_probes = pd.DataFrame(dict_probes)

    probes_array = []
    for index, row in df_probes.iterrows():
        if row['sonda'] != 'AAA' and df_probes.at[index, 'sonda'] not in probes_array:
            probes_array.append(df_probes.at[index, 'sonda'])
    #print(probes_array)

    if multiplex:
        dict_multiplex = multiplex_probes.multiplex_sequences(probes_array, mindg, maxdt)
        for index, row in df_probes.iterrows():
            if row['sonda'] == 'AAA':
                df_probes.at[index, 'grupo'] = -1
            else:
                df_probes.at[index, 'grupo'] = dict_multiplex[row['sonda']]

    return df_probes

progreso_compartido = 0
def actualizar_progreso(valor):
    global progreso_compartido
    progreso_compartido = valor

def generate_xlsx(df, name, minlen=60, maxlen=120, tmmin=65, tmmax=80, gcmin=30, gcmax=70, mindist=0, maxdist=50, minoverlap=25, maxoverlap=50, dgmin_homodim=-10000, dgmin_hairpin=-10000, maxhomopol_simple=6, maxhomopol_double=5, maxhomopol_triple=4, multiplex=True, mindg=-13627, maxdt=5):
    """
    Función que genera un reporte excel que contiene todas las sondas generadas y sus características. Además muestra las restricciones iniciales. 
    :param df: DataFrame de pandas que contiene las sondas y sus parámetros.
    :param name: Nombre que se le quiere dar al reporte.
    :param minlen: El largo mínimo de la sonda.
    :param maxlen: El largo máximo de la sonda.
    :param tmmin: Temperatura de melting mínima.
    :param tmmax: Temperatura de melting máxima.
    :param gcmin: Porcentaje de GC mínimo.
    :param gcmin: Porcentaje de GC máximo.
    :param mindist: Distancia mínima al borde del exón.
    :param maxdist: Distancia máxima al borde del exón.
    :param minoverlap: Sobrelape mínimo entre sondas.
    :param maxoverlap: Sobrelape máximo entre sondas.
    :param dgmin_homodim: Delta G mínimo permitido para validación de homodimerización.
    :param dgmin_hairpin: Delta G mínimo permitido para validación de hairpin u horquilla.
    :
    :return: Lista con pares de posiciones en una tupla: el fin de un exón y el inicio del siguiente.
    """
    
    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = 'Parámetros'

    sheet1.cell(row=1, column=1).value = 'PARÁMETROS DE DISEÑO'

    sheet1.cell(row=3, column=1).value = 'Largo de sonda mínimo'
    sheet1.cell(row=3, column=2).value = minlen

    sheet1.cell(row=4, column=1).value = 'Largo de sonda máximo'
    sheet1.cell(row=4, column=2).value = maxlen

    sheet1.cell(row=6, column=1).value = 'Temp Melting mínima'
    sheet1.cell(row=6, column=2).value = tmmin

    sheet1.cell(row=7, column=1).value = 'Temp Melting máxima'
    sheet1.cell(row=7, column=2).value = tmmax

    sheet1.cell(row=9, column=1).value = '%GC mínimo'
    sheet1.cell(row=9, column=2).value = gcmin

    sheet1.cell(row=10, column=1).value = '%GC máximo'
    sheet1.cell(row=10, column=2).value = gcmax

    sheet1.cell(row=12, column=1).value = 'Dist mínima al borde del exón'
    sheet1.cell(row=12, column=2).value = mindist
    
    sheet1.cell(row=13, column=1).value = 'Dist máxima al borde del exón'
    sheet1.cell(row=13, column=2).value = maxdist

    sheet1.cell(row=15, column=1).value = 'Sobrelape mínimo'
    sheet1.cell(row=15, column=2).value = minoverlap

    sheet1.cell(row=16, column=1).value = 'Sobrelape máximo'
    sheet1.cell(row=16, column=2).value = maxoverlap

    sheet1.cell(row=18, column=1).value = 'Delta G hairpin mínimo permitido'
    sheet1.cell(row=18, column=2).value = dgmin_hairpin

    sheet1.cell(row=19, column=1).value = 'Delta G homodimero mínimo permitido'
    sheet1.cell(row=19, column=2).value = dgmin_homodim

    sheet1.cell(row=21, column=1).value = 'Máximo homopolímeros simples permitidos'
    sheet1.cell(row=21, column=2).value = maxhomopol_simple

    sheet1.cell(row=22, column=1).value = 'Máximo homopolímeros dobles permitidos'
    sheet1.cell(row=22, column=2).value = maxhomopol_double

    sheet1.cell(row=23, column=1).value = 'Máximo homopolímeros triples permitidos'
    sheet1.cell(row=23, column=2).value = maxhomopol_triple

    if multiplex:
        sheet1.cell(row=25, column=1).value = 'Criterios de multiplexación'
        sheet1.cell(row=27, column=1).value = 'Diferencia máxima de Temp Melting'
        sheet1.cell(row=27, column=2).value = maxdt

        sheet1.cell(row=28, column=1).value = 'Mínimo delta G heterodimerización'
        sheet1.cell(row=28, column=2).value = mindg

    for cell in sheet1['A']:
        cell.style = 'Pandas'

    for column_cells in sheet1.columns:
        max_length = 2
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet1.column_dimensions[column].width = adjusted_width


    sheet2 = wb.create_sheet("Sondas")
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet2.append(r)

    for cell in sheet2['A'] + sheet2[1]:
        cell.style = 'Pandas'

    #Ajustar ancho de columnas
    for column_cells in sheet2.columns:
        max_length = 2
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet2.column_dimensions[column].width = adjusted_width

    
    now = datetime.now()
    filename = name+'_'+now.strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(os.getcwd(),'sondas', filename, filename+'.xlsx')

    folder_path = os.path.join(os.getcwd(), "sondas", filename)
    if not os.path.exists(folder_path):
        os.mkdir(folder_path)

    wb.save(filepath)
    return filename

def get_all_transcripts(seqrecord):
    transcripts = []
    for f in seqrecord.features:
        if f.type == 'CDS':
            info = f.qualifiers.get('product')[0]
            if str(info) not in transcripts:
                transcripts.append(str(info))
    return transcripts

def get_all_genes(seqrecord):
    genes = []
    for f in seqrecord.features:
        if f.type == 'CDS':
            gen = f.qualifiers.get('gene')[0]
            if str(gen) not in genes:
                genes.append(str(gen))
    return genes


def main():
    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-an',
                        '--accessionnumber',
                        help='Accesion number de la secuencia a analizar.',
                        #required=True,
                        default=argparse.SUPPRESS)
    parser.add_argument('-minlen', '--minimumlength',help='Largo mínimo de sonda.' ,type=int,default=60)
    parser.add_argument('-maxlen', '--maximumlength',help='Largo mínimo de sonda.' ,type=int,default=120)
    parser.add_argument('-tmmin', '--mintempmelting',help='Temperatura de melting mínima.' ,type=int,default=65)
    parser.add_argument('-tmmax', '--maxtempmelting',help='Temperatura de melting máxima.' ,type=int,default=80)
    parser.add_argument('-gcmin', '--minumumgc',help='Porcentaje de GC mínimo.' ,type=int,default=30)
    parser.add_argument('-gcmax', '--maximumgc',help='Porcentaje de GC máximo.' ,type=int,default=70)
    parser.add_argument('-olmin', '--minimumoverlap',help='Sobrelape mínimo.' ,type=int,default=25)
    parser.add_argument('-olmax', '--maximumoverlap',help='Sobrelape máximo.' ,type=int,default=50)
    parser.add_argument('-dghairpin', '--deltaghairpin',help='Mínimo Delta G hairpin permitido.' ,type=int,default=-8000)
    parser.add_argument('-dghomodim', '--deltaghomodimer',help='Mínimo Delta G homodimero permitido.' ,type=int,default=-8000)

    args = parser.parse_args()
    dargs = vars(args)
    rec = descarga.parse_file_to_seqrecord('C:/Users/simon/Downloads/tp53.gb')
    #rec = descarga.accnum_to_seqrecord(dargs["accessionnumber"])
    df = probe_designer(record=rec, transcripts=get_all_transcripts(rec))
                        # ,
                        # minlen=dargs["minimumlength"],
                        # maxlen=dargs["maximumlength"],
                        # tmmin=dargs["mintempmelting"],
                        # tmmax=dargs["maxtempmelting"],
                        # gcmin=dargs["minumumgc"],
                        # gcmax=dargs["maximumgc"],
                        # minoverlap=dargs["minimumoverlap"],
                        # maxoverlap=dargs["maximumoverlap"],
                        # maxoverlap=dargs["deltaghairpin"],
                        # maxoverlap=dargs["deltaghomodimer"])
    
    generate_xlsx(df=df,
                  name='sondas',
                  minlen=dargs["minimumlength"],
                  maxlen=dargs["maximumlength"],
                  tmmin=dargs["mintempmelting"],
                  tmmax=dargs["maxtempmelting"],
                  gcmin=dargs["minumumgc"],
                  gcmax=dargs["maximumgc"],
                  minoverlap=dargs["minimumoverlap"],
                  maxoverlap=dargs["maximumoverlap"])

if __name__ == "__main__":
    main()