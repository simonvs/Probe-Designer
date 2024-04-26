import descarga
import multiplex
import os
import argparse
import pandas as pd
import primer3
import subprocess
import queue
try:
    from Bio.SeqUtils import gc_fraction
    def GC(sequence):
        return 100 * gc_fraction(sequence, ambiguous="ignore")
except ImportError:
    # Older versions have this:
    from Bio.SeqUtils import GC
from Bio.SeqUtils import MeltingTemp
from Bio.Seq import Seq
from Bio import SearchIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


def check_probe(seq, iscentral, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple):
    """
    Función que recibe una sonda (secuencia de nucleótidos) y las restricciones para verificar que la
    sonda cumpla los parámetros establecidos.
    Retorna un valor booleano que determina si cumple las restricciones o no.

    :param seq: La secuencia que se quiere verificar.
    :param iscentral: Booleano que indica si la sonda es central (una mitad en un exón y la otra en el siguiente)
    :param minlen: El largo mínimo de la sonda.
    :param maxlen: El largo máximo de la sonda.
    :param tmmin: Temperatura de melting mínima.
    :param tmmax: Temperatura de melting máxima.
    :param gcmin: Porcentaje de GC mínimo.
    :param gcmax: Porcentaje de GC máximo.
    :param dgmin_homodim: Delta G mínimo permitido para validación de homodimerización.
    :param dgmin_hairpin: Delta G mínimo permitido para validación de hairpin u horquilla.
    :param maxhomopol_simple: Cantidad máxima permitida de homopolímeros simples (AAAAA)
    :param maxhomopol_double: Cantidad máxima permitida de homopolímeros dobles (AGAGAGAG)
    :param maxhomopol_triple: Cantidad máxima permitida de homopolímeros triples (AGCAGCAGC)
    :return: Booleano que determina la validez de la secuencia
    """

    #Chequear largo
    largo = len(seq)
    if largo < minlen or largo > maxlen:
        #print('Largo no valido')
        return False
    
    #Chequear Tm
    tm = MeltingTemp.Tm_NN(seq)
    #tm = primer3.calc_tm(str(seq))
    if tm < tmmin or tm > tmmax:
        #print('Tm fuera de rango')
        return False
    
    #Chequear %GC
    if GC(seq) < gcmin or GC(seq) > gcmax:
        #print('GC fuera de rango')
        return False
    
    #Chequear hetero - homodimeros - horquilla
    #dg_hairpin = primer3.calcHairpin(str(seq)).dg
    if largo <= 60:
        if primer3.calcHairpin(str(seq)).dg < dgmin_hairpin:
            #print('Delta G hairpin')
            return False
        if primer3.calcHomodimer(str(seq)).dg < dgmin_homodim:
            #print('Delta G homodim')
            return False
    else:
        if primer3.calcHairpin(str(seq[:60])).dg < dgmin_hairpin or primer3.calcHairpin(str(seq[60:])).dg < dgmin_hairpin:
            #print('Delta G hairpin')
            return False
        if primer3.calcHomodimer(str(seq[:60])).dg < dgmin_homodim or primer3.calcHomodimer(str(seq[60:])).dg < dgmin_homodim:
            #print('Delta G homodim')
            return False

    
    # Chequear homopolimeros de 1 base
    bases = ['A', 'C', 'G', 'T']
    for b in bases:
        if b * (maxhomopol_simple + 1) in seq:
            #print('Homopolimero de 1 nucleotido')
            return False
        
    # Chequear homopolimeros de 2 bases
    checked = []
    for i in range(len(seq)-2):
        subseq = str(seq[i:i+2])
        if subseq not in checked:
            if subseq * (maxhomopol_double + 1) in seq:
                #print('Homopolimero de 2 nucleotidos')
                return False
            checked.append(subseq)

    # Chequear homopolimeros de 3 bases
    checked = []
    for i in range(len(seq)-3):
        subseq = str(seq[i:i+3])
        if subseq not in checked:
            if subseq * (maxhomopol_triple + 1) in seq:
                #print('Homopolimero de 3 nucleotidos')
                return False
            checked.append(subseq)
    
    #Chequear especificidad (BLAST LOCAL)
    #if not verify_specificity(seq, iscentral):
    #    print('Especificidad')
    #    return False
    
    print('Sonda Aceptada: ', str(seq))
    return True

def get_probes_from_pos(empalme, record, site, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, mindist, maxdist, minoverlap, maxoverlap, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple):
    """
    Esta función genera un par de sondas a partir de la secuencia de referencia y dos posiciones que representan el punto de empalme entre dos exones.
    El par de sondas generadas pueden ubicarse en el exón donor, acceptor o en ambos (central).

    :param empalme: Tupla con posiciones que representan al punto de empalme o splicing.
    :param record: Secuencia de referencia.
    :param site: Entero que representa si las sondas son donor, acceptor o central.
    :param minlen: El largo mínimo de la sonda.
    :param maxlen: El largo máximo de la sonda.
    :param tmmin: Temperatura de melting mínima.
    :param tmmax: Temperatura de melting máxima.
    :param gcmin: Porcentaje de GC mínimo.
    :param gcmin: Porcentaje de GC máximo.
    :param mindist: Distancia mínima al borde del exón.
    :param maxdist: Distancia máxima al borde del exón.
    :param minoverlap: Sobrelape mínimo entre sondas.
    :param maxoverlap: Sobrelape máximo entre sondas.
    :param dgmin_homodim: Delta G mínimo permitido para validación de homodimerización.
    :param dgmin_hairpin: Delta G mínimo permitido para validación de hairpin u horquilla.
    :param maxhomopol_simple: Cantidad máxima permitida de homopolímeros simples (AAAAA)
    :param maxhomopol_double: Cantidad máxima permitida de homopolímeros dobles (AGAGAGAG)
    :param maxhomopol_triple: Cantidad máxima permitida de homopolímeros triples (AGCAGCAGC)
    :return: Lista con dos sondas, cada sonda es una tupla con la secuencia, la posición de inicio, la posición final y la distancia al borde del exón si aplica.
    """
    b = False
    if site == 0:
        i = minlen
        while i <= maxlen and not b:
            pos_inicio_int = empalme[0]-i//2
            pos_fin_int = empalme[1]+i//2
            probe_int = (record.seq[pos_inicio_int:empalme[0]] + record.seq[empalme[1]:pos_fin_int]).reverse_complement()
            b = check_probe(probe_int, True, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
            i += 1
        if not b:
            return [(Seq('AAA'), -1, -1, -1), (Seq('AAA'), -1, -1, -1)]
        else:
            b = False
            i += 3
            while i <= maxlen and not b:
                pos_inicio_ext = empalme[0]-i//2
                pos_fin_ext = empalme[1]+i//2
                probe_ext = (record.seq[pos_inicio_ext:empalme[0]] + record.seq[empalme[1]:pos_fin_ext]).reverse_complement()
                b = check_probe(probe_ext, True, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                i += 1
            if b:
                return [(probe_int, pos_inicio_int, pos_fin_int, -1), (probe_ext, pos_inicio_ext, pos_fin_ext, -1)]
            else:
                return [(probe_int, pos_inicio_int, pos_fin_int, -1), (Seq('AAA'), -1, -1, -1)]


    elif site == -1:
        dist_to_border = mindist
        while dist_to_border <= maxdist and not b:
            i = minlen
            while i <= maxlen and not b:
                pos_inicio_int = empalme[0]-i-dist_to_border
                pos_fin_int = empalme[0]-dist_to_border
                probe_int = record.seq[pos_inicio_int:pos_fin_int].reverse_complement()
                b = check_probe(probe_int, False, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                i += 1
            dist_to_border += 1
        if not b:
            return [(Seq('AAA'), -1, -1, -1), (Seq('AAA'), -1, -1, -1)]    
        else:
            b = False
            dist_int_probe = dist_to_border
            dist_to_border += i * (100-maxoverlap) / 100
            while dist_to_border <= dist_int_probe + i * (100-minoverlap) / 100 and not b:
                j = minlen
                while j <= maxlen and not b:
                    pos_inicio_ext = int(empalme[0]-j-dist_to_border)
                    pos_fin_ext = int(empalme[0]-dist_to_border)
                    probe_ext = record.seq[pos_inicio_ext:pos_fin_ext].reverse_complement()
                    b = check_probe(probe_ext, False, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                    j += 1
                dist_to_border += 1
            if b:
                return [(probe_int, pos_inicio_int, pos_fin_int, dist_int_probe), (probe_ext, pos_inicio_ext, pos_fin_ext, dist_to_border)]
            else:
                return [(probe_int, pos_inicio_int, pos_fin_int, dist_int_probe), (Seq('AAA'), -1, -1, -1)]
            

    elif site == 1:
        dist_to_border = mindist
        while dist_to_border <= maxdist and not b:
            i = minlen
            while i <= maxlen and not b:
                pos_inicio_int = empalme[1]+dist_to_border
                pos_fin_int = empalme[1]+i+dist_to_border
                probe_int = record.seq[pos_inicio_int:pos_fin_int].reverse_complement()
                b = check_probe(probe_int, False, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                i += 1
            dist_to_border += 1
        
        if not b:
            return [(Seq('AAA'), -1, -1, -1), (Seq('AAA'), -1, -1, -1)]    
        else:
            b = False
            dist_int_probe = dist_to_border
            dist_to_border += i * (100-maxoverlap) / 100
            while dist_to_border <= dist_int_probe + i * (100-minoverlap) / 100 and not b:
                j = minlen
                while j <= maxlen and not b:
                    pos_inicio_ext = int(empalme[1]+dist_to_border)
                    pos_fin_ext = int(empalme[1]+j+dist_to_border)
                    probe_ext = record.seq[pos_inicio_ext:pos_fin_ext].reverse_complement()
                    b = check_probe(probe_ext, False, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)
                    j += 1
                dist_to_border += 1
            if b:
                return [(probe_int, pos_inicio_int, pos_fin_int, dist_int_probe), (probe_ext, pos_inicio_ext, pos_fin_ext, dist_to_border)]
            else:
                return [(probe_int, pos_inicio_int, pos_fin_int, dist_int_probe), (Seq('AAA'), -1, -1, -1)]

def verify_specificity(seq, iscentral):
    """
    Función que verifica la especificidad o unicidad de una secuencia usando la herramienta de alineamiento BLAST de manera local.
    Es necesario tener instalado Blast+ y las bases de datos en la carpeta 'refseq'.
    
    :param seq: Secuencia de nuecleótidos que se quiere validar.
    :param iscentral: Booleano que indica si se trata de una sonda central. Se utilizan diferentes criterios para chequear la especificidad de una sonda central.
    :return: Booleano que indica si la secuencia se considera específica.
    """
    db1_name = "refseq/db1_genoma"
    db2_name = "refseq/db2_transcriptoma"
    query = "refseq/query.txt"
    outxml = 'refseq/resultados.xml'

    with open(query, "w") as archivo:
        archivo.write(">1\n")
        archivo.write(str(seq))

    blastn_cmd = ['blastn',
                '-task', 'blastn-short',
                '-query', query,
                '-db', db2_name,
                '-out', outxml,
                '-outfmt', '5',
                '-word_size', '11',
                '-dust', 'no']

    subprocess.run(blastn_cmd, shell=True)
    blast_qresult = SearchIO.read("refseq/resultados.xml", "blast-xml")

    if iscentral:
        if blast_qresult[0].hsps[2].evalue / blast_qresult[0].hsps[1].evalue < 100:
            return False
        if blast_qresult[1].hsps[0].evalue / blast_qresult[0].hsps[1].evalue < 100:
            return False
        
    else:
        if blast_qresult[0].hsps[1].evalue / blast_qresult[0].hsps[0].evalue < 100:
            return False
        if blast_qresult[1].hsps[0].evalue / blast_qresult[0].hsps[0].evalue < 100:
            return False
    return True

def get_splicing_pairs(locations):
    """
    Función que obtiene las zonas de empalme a partir de una serie de regiones que representan una transcripción.
    Las zonas de empalme se representan por dos posiciones: el fin de un exón y el inicio del siguiente.
    
    :param locations: Subregiones que representan la secuencia codificante o secuencia de exones
    :return: Lista con pares de posiciones en una tupla: el fin de un exón y el inicio del siguiente.
    """
    positions = []
    for subloc in locations.parts:
        positions.append((subloc.start, subloc.end))

    empalmes = []
    for i in range(len(positions)-1):
        empalmes.append((positions[i][1].position, positions[i+1][0].position))
    return empalmes

def get_splicings(seqrecord, transcripts):
    """
    Función que retorna los empalmes sin repetición de una secuencia anotada en ciertos transcritos.
    
    :param seqrecord: Secuencia anotada de referencia
    :param transcripts: Lista de transcritos
    :return: Lista con pares de posiciones en una tupla: el fin de un exón y el inicio del siguiente.
    """
    empalmes_array = []
    for f in seqrecord.features:
        if f.type == 'CDS':
            if f.qualifiers.get('product')[0] in transcripts:
                loc = f.location
                pares_empalme = get_splicing_pairs(loc)
                for par in pares_empalme:
                    if par not in empalmes_array:
                        empalmes_array.append(par)
    return empalmes_array

def probe_designer(record, transcripts, progreso_queue=None, minlen=60, maxlen=120, tmmin=65, tmmax=80, gcmin=30, gcmax=70, mindist=0, maxdist=200, minoverlap=25, maxoverlap=50, dgmin_homodim=-10000, dgmin_hairpin=-10000, maxhomopol_simple=6, maxhomopol_double=5, maxhomopol_triple=4, ismultiplex=True, mindg=-13627, maxdt=5):
    """
    Función principal diseñadora de sondas. De la secuencia anotada se obtiene una serie de sondas como subsecuencias de ésta, que cumplen varias restricciones ajustadas por los parámetros.
    
    :param record: Secuencia anotada de referencia para la creación de las sondas.
    :param transcripts: Transcripciones en las cuales se quiere diseñar.
    :param minlen: El largo mínimo de la sonda.
    :param maxlen: El largo máximo de la sonda.
    :param tmmin: Temperatura de melting mínima.
    :param tmmax: Temperatura de melting máxima.
    :param gcmin: Porcentaje de GC mínimo.
    :param gcmin: Porcentaje de GC máximo.
    :param mindist: Distancia mínima al borde del exón.
    :param maxdist: Distancia máxima al borde del exón.
    :param minoverlap: Sobrelape mínimo entre sondas.
    :param maxoverlap: Sobrelape máximo entre sondas.
    :param dgmin_homodim: Delta G mínimo permitido para validación de homodimerización.
    :param dgmin_hairpin: Delta G mínimo permitido para validación de hairpin u horquilla.
    :param maxhomopol_simple: Cantidad máxima permitida de homopolímeros simples (AAAAA)
    :param maxhomopol_double: Cantidad máxima permitida de homopolímeros dobles (AGAGAGAG)
    :param maxhomopol_triple: Cantidad máxima permitida de homopolímeros triples (AGCAGCAGC)
    :param ismultiplex: Booleano que indica si el diseño considera multiplexación
    :param mindg: Delta G mínimo para heterodimerización (multiplex)
    :param maxdt: Diferencia máxima de Tm (multiplex)
    :return: Dataframe con las sondas y sus características
    """
    locations = []
    dict_cds = {
        'gen':[],
        'info':[],
        'empalmes':[]
    }

    for f in record.features:
        if f.type == 'CDS':
            loc = f.location
            info = f.qualifiers.get('product', ['Transctipto no identificado'])[0]
            if str(loc) not in locations and info in transcripts:
                locations.append(str(loc))
                dict_cds['gen'].append(f.qualifiers.get('gene', ['Gen no identificado'])[0])
                dict_cds['info'].append(info)
                dict_cds['empalmes'].append(get_splicing_pairs(loc))

    df_cds = pd.DataFrame(dict_cds)

    dict_probes = {
        'nombre':[],
        'gen':[],
        'transcrito':[],
        'sonda':[],
        'pos inicio':[],
        'pos fin':[],
        'empalme':[],
        'lado':[],
        'tm':[],
        'gc':[],
        'largo':[],
        'dist':[]
        #'deltag hairpin':[],
        #'deltag homodim':[]
    }

    splicing_pairs = get_splicings(record, transcripts)
    dict_generated_probes = {}
    num_splicings = len(splicing_pairs)

    progreso_actual = 0

    for index, row in df_cds.iterrows():
        for empalme in row['empalmes']:
            #Verificar si aún no se diseñan las sondas para ese empalme
            if str(empalme) not in dict_generated_probes:
                donor_probes = get_probes_from_pos(empalme, record, -1, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, mindist, maxdist, minoverlap, maxoverlap, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)

                #Sonda donor interior
                dict_probes['nombre'].append('DI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])               
                dict_probes['sonda'].append(str(donor_probes[0][0]))
                dict_probes['pos inicio'].append(donor_probes[0][1])
                dict_probes['pos fin'].append(donor_probes[0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('donor interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(donor_probes[0][0]))
                dict_probes['gc'].append(int(GC(donor_probes[0][0])))
                dict_probes['largo'].append(len(donor_probes[0][0]))
                dict_probes['dist'].append(int(donor_probes[0][3]))

                #Sonda donor exterior
                dict_probes['nombre'].append('DE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])                
                dict_probes['sonda'].append(str(donor_probes[1][0]))
                dict_probes['pos inicio'].append(donor_probes[1][1])
                dict_probes['pos fin'].append(donor_probes[1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('donor exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(donor_probes[1][0]))
                dict_probes['gc'].append(int(GC(donor_probes[1][0])))
                dict_probes['largo'].append(len(donor_probes[1][0]))
                dict_probes['dist'].append(int(donor_probes[1][3]))

                acceptor_probes = get_probes_from_pos(empalme, record, 1, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, mindist, maxdist, minoverlap, maxoverlap, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)

                #Sonda acceptor interior
                dict_probes['nombre'].append('AI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])                
                dict_probes['sonda'].append(str(acceptor_probes[0][0]))
                dict_probes['pos inicio'].append(acceptor_probes[0][1])
                dict_probes['pos fin'].append(acceptor_probes[0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('acceptor interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(acceptor_probes[0][0]))
                dict_probes['gc'].append(int(GC(acceptor_probes[0][0])))
                dict_probes['largo'].append(len(acceptor_probes[0][0]))
                dict_probes['dist'].append(int(acceptor_probes[0][3]))


                #Sonda acceptor exterior
                dict_probes['nombre'].append('AE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(acceptor_probes[1][0]))
                dict_probes['pos inicio'].append(acceptor_probes[1][1])
                dict_probes['pos fin'].append(acceptor_probes[1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('acceptor exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(acceptor_probes[1][0]))
                dict_probes['gc'].append(int(GC(acceptor_probes[1][0])))
                dict_probes['largo'].append(len(acceptor_probes[1][0]))
                dict_probes['dist'].append(int(acceptor_probes[1][3]))

                central_probes = get_probes_from_pos(empalme, record, 0, minlen, maxlen, tmmin, tmmax, gcmin, gcmax, mindist, maxdist, minoverlap, maxoverlap, dgmin_homodim, dgmin_hairpin, maxhomopol_simple, maxhomopol_double, maxhomopol_triple)

                #Sonda central interior
                dict_probes['nombre'].append('CI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])    
                dict_probes['sonda'].append(str(central_probes[0][0]))
                dict_probes['pos inicio'].append(central_probes[0][1])
                dict_probes['pos fin'].append(central_probes[0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('central interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(central_probes[0][0]))
                dict_probes['gc'].append(int(GC(central_probes[0][0])))
                dict_probes['largo'].append(len(central_probes[0][0]))
                dict_probes['dist'].append(int(central_probes[0][3]))

                #Sonda central exterior
                dict_probes['nombre'].append('CE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])                
                dict_probes['sonda'].append(str(central_probes[1][0]))
                dict_probes['pos inicio'].append(central_probes[1][1])
                dict_probes['pos fin'].append(central_probes[1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('central exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(central_probes[1][0]))
                dict_probes['gc'].append(int(GC(central_probes[1][0])))
                dict_probes['largo'].append(len(central_probes[1][0]))
                dict_probes['dist'].append(int(central_probes[1][3]))

                dict_generated_probes[str(empalme)] = [donor_probes, acceptor_probes, central_probes]

                progreso_actual += 100 // num_splicings

                if (progreso_queue):
                    progreso_queue.put(progreso_actual)
                #print(progreso_actual)


            #Si el punto de empalme ya fue diseñado se extrae del diccionario
            else:
                #Sonda donor interior
                dict_probes['nombre'].append('DI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])                
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][0][0][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][0][0][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][0][0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('donor interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][0][0][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][0][0][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][0][0][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][0][0][3]))

                #Sonda donor exterior
                dict_probes['nombre'].append('DE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])               
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][0][1][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][0][1][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][0][1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('donor exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][0][1][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][0][1][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][0][1][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][0][1][3]))


                #Sonda acceptor interior
                dict_probes['nombre'].append('AI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][1][0][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][1][0][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][1][0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('acceptor interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][1][0][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][1][0][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][1][0][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][1][0][3]))

                #Sonda acceptor exterior
                dict_probes['nombre'].append('AE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][1][1][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][1][1][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][1][1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('acceptor exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][1][1][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][1][1][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][1][1][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][1][1][3]))


                #Sonda central interior
                dict_probes['nombre'].append('CI'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][2][0][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][2][0][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][2][0][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('central interior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][2][0][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][2][0][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][2][0][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][2][0][3]))


                #Sonda central exterior
                dict_probes['nombre'].append('CE'+str(splicing_pairs.index(empalme)+1))
                dict_probes['gen'].append(row['gen'])
                dict_probes['transcrito'].append(row['info'])
                dict_probes['sonda'].append(str(dict_generated_probes[str(empalme)][2][1][0]))
                dict_probes['pos inicio'].append(dict_generated_probes[str(empalme)][2][1][1])
                dict_probes['pos fin'].append(dict_generated_probes[str(empalme)][2][1][2])
                dict_probes['empalme'].append(str(empalme))
                dict_probes['lado'].append('central exterior')
                dict_probes['tm'].append(MeltingTemp.Tm_NN(dict_generated_probes[str(empalme)][2][1][0]))
                dict_probes['gc'].append(int(GC(dict_generated_probes[str(empalme)][2][1][0])))
                dict_probes['largo'].append(len(dict_generated_probes[str(empalme)][2][1][0]))
                dict_probes['dist'].append(int(dict_generated_probes[str(empalme)][2][1][3]))


    df_probes = pd.DataFrame(dict_probes)

    probes_array = []
    for index, row in df_probes.iterrows():
        if row['sonda'] != 'AAA' and df_probes.at[index, 'sonda'] not in probes_array:
            probes_array.append(df_probes.at[index, 'sonda'])
    #print(probes_array)

    if ismultiplex:
        dict_multiplex = multiplex.multiplex_sequences(probes_array, progreso_queue, mindg, maxdt)
        for index, row in df_probes.iterrows():
            if row['sonda'] == 'AAA':
                df_probes.at[index, 'grupo'] = -1
            else:
                df_probes.at[index, 'grupo'] = dict_multiplex[row['sonda']]

    return df_probes

def generate_xlsx(df, name, genes, minlen=60, maxlen=120, tmmin=65, tmmax=80, gcmin=30, gcmax=70, mindist=0, maxdist=50, minoverlap=25, maxoverlap=50, dgmin_homodim=-10000, dgmin_hairpin=-10000, maxhomopol_simple=6, maxhomopol_double=5, maxhomopol_triple=4, multiplex=True, mindg=-13627, maxdt=5, tiempo_diseno=0):
    """
    Función que genera un reporte excel que contiene todas las sondas generadas y sus características. Además muestra las restricciones iniciales y los grupos de multiplexación.
    Almacena el reporte en la carpeta 'sondas' 
    
    :param df: DataFrame de pandas que contiene las sondas y sus parámetros.
    :param name: Nombre que se le quiere dar al reporte.
    :param minlen: El largo mínimo de la sonda.
    :param maxlen: El largo máximo de la sonda.
    :param tmmin: Temperatura de melting mínima.
    :param tmmax: Temperatura de melting máxima.
    :param gcmin: Porcentaje de GC mínimo.
    :param gcmin: Porcentaje de GC máximo.
    :param mindist: Distancia mínima al borde del exón.
    :param maxdist: Distancia máxima al borde del exón.
    :param minoverlap: Sobrelape mínimo entre sondas.
    :param maxoverlap: Sobrelape máximo entre sondas.
    :param dgmin_homodim: Delta G mínimo permitido para validación de homodimerización.
    :param dgmin_hairpin: Delta G mínimo permitido para validación de hairpin u horquilla.
    :param maxhomopol_simple: Cantidad máxima permitida de homopolímeros simples (AAAAA)
    :param maxhomopol_double: Cantidad máxima permitida de homopolímeros dobles (AGAGAGAG)
    :param maxhomopol_triple: Cantidad máxima permitida de homopolímeros triples (AGCAGCAGC)
    :param multiplex: Booleano que indica si el diseño considera multiplexación
    :param mindg: Delta G mínimo para heterodimerización (multiplex)
    :param maxdt: Diferencia máxima de Tm (multiplex)
    :param tiempo_diseno: Tiempo que demoró la ejecución en segundos
    :return: Nombre del archivo XLSX generado
    """
    if not os.path.exists(os.path.join(os.getcwd(),'sondas')):
        os.makedirs(os.path.join(os.getcwd(),'sondas'))
        
    wb = Workbook()

    df_resumen = df.drop_duplicates(subset='sonda')
    df_resumen = df_resumen[df_resumen['sonda'] != "AAA"]

    sheet1 = wb.active
    sheet1.title = 'Resumen'

    sheet1.cell(row=1, column=1).value = 'Genes: '+str(genes)

    sheet1.cell(row=3, column=1).value = 'Número deseable de sondas'
    sheet1.cell(row=3, column=2).value = df.shape[0]

    sheet1.cell(row=4, column=1).value = 'Número de sondas válidas'
    sheet1.cell(row=4, column=2).value = df[df['sonda'] != "AAA"].shape[0]

    sheet1.cell(row=5, column=1).value = 'Número de sondas diferentes'
    sheet1.cell(row=5, column=2).value = df_resumen.shape[0]

    sheet1.cell(row=7, column=1).value = 'Tm promedio'
    sheet1.cell(row=7, column=2).value = str("{:.1f}").format(df_resumen['tm'].mean())+'°'

    sheet1.cell(row=8, column=1).value = '%GC promedio'
    sheet1.cell(row=8, column=2).value = str("{:.1f}").format(df_resumen['gc'].mean())+'%'

    sheet1.cell(row=9, column=1).value = 'Largo promedio'
    sheet1.cell(row=9, column=2).value = str("{:.1f}").format(df_resumen['largo'].mean())+' nt'

    if multiplex:
        sheet1.cell(row=13, column=1).value = 'Número de grupos (multiplex)'
        sheet1.cell(row=13, column=2).value = int(df_resumen['grupo'].max())


    h1, m1, s1 = segundos_a_hms(tiempo_diseno)
    sheet1.cell(row=11, column=1).value = 'Tiempo de ejecución'
    sheet1.cell(row=11, column=2).value = str(h1)+':'+str(m1)+':'+str(s1)

    for cell in sheet1['A']:
        cell.style = 'Pandas'
    for cell in sheet1['B']:
        cell.style = 'Pandas'


    for column_cells in sheet1.columns:
        max_length = 2
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet1.column_dimensions[column].width = adjusted_width

    sheet2 = wb.create_sheet("Parámetros iniciales")

    sheet2.cell(row=1, column=1).value = 'Parámetros iniciales'

    sheet2.cell(row=3, column=1).value = 'Largo de sonda mínimo'
    sheet2.cell(row=3, column=2).value = minlen

    sheet2.cell(row=4, column=1).value = 'Largo de sonda máximo'
    sheet2.cell(row=4, column=2).value = maxlen

    sheet2.cell(row=6, column=1).value = 'Temp Melting mínima'
    sheet2.cell(row=6, column=2).value = tmmin

    sheet2.cell(row=7, column=1).value = 'Temp Melting máxima'
    sheet2.cell(row=7, column=2).value = tmmax

    sheet2.cell(row=9, column=1).value = '%GC mínimo'
    sheet2.cell(row=9, column=2).value = gcmin

    sheet2.cell(row=10, column=1).value = '%GC máximo'
    sheet2.cell(row=10, column=2).value = gcmax

    sheet2.cell(row=12, column=1).value = 'Dist mínima al borde del exón'
    sheet2.cell(row=12, column=2).value = mindist
    
    sheet2.cell(row=13, column=1).value = 'Dist máxima al borde del exón'
    sheet2.cell(row=13, column=2).value = maxdist

    sheet2.cell(row=15, column=1).value = 'Sobrelape mínimo'
    sheet2.cell(row=15, column=2).value = minoverlap

    sheet2.cell(row=16, column=1).value = 'Sobrelape máximo'
    sheet2.cell(row=16, column=2).value = maxoverlap

    sheet2.cell(row=18, column=1).value = 'Delta G hairpin mínimo permitido'
    sheet2.cell(row=18, column=2).value = dgmin_hairpin

    sheet2.cell(row=19, column=1).value = 'Delta G homodimero mínimo permitido'
    sheet2.cell(row=19, column=2).value = dgmin_homodim

    sheet2.cell(row=21, column=1).value = 'Máximo homopolímeros simples permitidos'
    sheet2.cell(row=21, column=2).value = maxhomopol_simple

    sheet2.cell(row=22, column=1).value = 'Máximo homopolímeros dobles permitidos'
    sheet2.cell(row=22, column=2).value = maxhomopol_double

    sheet2.cell(row=23, column=1).value = 'Máximo homopolímeros triples permitidos'
    sheet2.cell(row=23, column=2).value = maxhomopol_triple

    if multiplex:
        sheet2.cell(row=25, column=1).value = 'Criterios de multiplexación'
        sheet2.cell(row=27, column=1).value = 'Diferencia máxima de Temp Melting'
        sheet2.cell(row=27, column=2).value = maxdt

        sheet2.cell(row=28, column=1).value = 'Mínimo delta G heterodimerización'
        sheet2.cell(row=28, column=2).value = mindg

    for cell in sheet2['A']:
        cell.style = 'Pandas'
    for cell in sheet2['B']:
        cell.style = 'Pandas'

    for column_cells in sheet2.columns:
        max_length = 2
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet2.column_dimensions[column].width = adjusted_width


    sheet3 = wb.create_sheet("Grupos")

    sheet3.cell(row=1, column=1).value = 'GRUPO'  
    sheet3.cell(row=1, column=2).value = 'Cantidad de sondas'
    sheet3.cell(row=1, column=3).value = 'Tm promedio'
    sheet3.cell(row=1, column=4).value = '%GC promedio'
    sheet3.cell(row=1, column=5).value = 'Largo promedio'

    if multiplex:
        for i in range(int(df_resumen['grupo'].max())):
            grupo = i+1
            sheet3.cell(row=1+grupo, column=1).value = grupo
            sheet3.cell(row=1+grupo, column=2).value = df_resumen[df_resumen['grupo'] == grupo]['grupo'].count()
            sheet3.cell(row=1+grupo, column=3).value = str("{:.1f}").format(df_resumen[df_resumen['grupo'] == grupo]['tm'].mean())
            sheet3.cell(row=1+grupo, column=4).value = str("{:.1f}").format(df_resumen[df_resumen['grupo'] == grupo]['gc'].mean())
            sheet3.cell(row=1+grupo, column=5).value = str("{:.1f}").format(df_resumen[df_resumen['grupo'] == grupo]['largo'].mean())

    for cell in sheet3['A']:
        cell.style = 'Pandas'
    for cell in sheet3['B']:
        cell.style = 'Pandas'
    for cell in sheet3['C']:
        cell.style = 'Pandas'
    for cell in sheet3['D']:
        cell.style = 'Pandas'
    for cell in sheet3['E']:
        cell.style = 'Pandas'

    for column_cells in sheet3.columns:
        max_length = 2
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet3.column_dimensions[column].width = adjusted_width

    sheet4 = wb.create_sheet("Sondas")
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet4.append(r)

    for cell in sheet4['A'] + sheet4[1]:
        cell.style = 'Pandas'

    #Ajustar ancho de columnas
    for column_cells in sheet4.columns:
        max_length = 2
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet4.column_dimensions[column].width = adjusted_width


    sheet5 = wb.create_sheet("Sondas filtradas")
    for r in dataframe_to_rows(df_resumen, index=False, header=True):
        sheet5.append(r)

    for cell in sheet5['A'] + sheet5[1]:
        cell.style = 'Pandas'

    #Ajustar ancho de columnas
    for column_cells in sheet5.columns:
        max_length = 2
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet5.column_dimensions[column].width = adjusted_width

    now = datetime.now()
    filename = name+'_'+now.strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(os.getcwd(),'sondas', filename, name+'.xlsx')

    folder_path = os.path.join(os.getcwd(), "sondas", filename)
    if not os.path.exists(folder_path):
        os.mkdir(folder_path)

    wb.save(filepath)
    return filename

def get_all_transcripts(seqrecord):
    """
    Función que retorna todos los transcritos anotados dentro de una secuancia.
    
    :param seqrecord: Subregiones que representan la secuencia codificante o secuencia de exones
    :return: Lista con pares de posiciones en una tupla: el fin de un exón y el inicio del siguiente.
    """
    transcripts = []
    for f in seqrecord.features:
        if f.type == 'CDS':
            info = f.qualifiers.get('product')[0]
            if str(info) not in transcripts:
                transcripts.append(str(info))
    return transcripts

def get_all_genes(seqrecord):
    """
    Obtener todos los genes anotados 
    
    :param seqrecord: Subregiones que representan la secuencia codificante o secuencia de exones
    :return: Lista con pares de posiciones en una tupla: el fin de un exón y el inicio del siguiente.
    """
    genes = []
    for f in seqrecord.features:
        if f.type == 'CDS':
            gen = f.qualifiers.get('gene')[0]
            if str(gen) not in genes:
                genes.append(str(gen))
    return genes

def segundos_a_hms(segundos):
    horas, segundos = divmod(segundos, 3600)
    minutos, segundos = divmod(segundos, 60)
    return horas, minutos, segundos

def main():
    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-an',
                        '--accessionnumber',
                        help='Accesion number de la secuencia a analizar.',
                        #required=True,
                        default=argparse.SUPPRESS)
    parser.add_argument('-minlen', '--minimumlength',help='Largo mínimo de sonda.' ,type=int,default=60)
    parser.add_argument('-maxlen', '--maximumlength',help='Largo mínimo de sonda.' ,type=int,default=120)
    parser.add_argument('-tmmin', '--mintempmelting',help='Temperatura de melting mínima.' ,type=int,default=65)
    parser.add_argument('-tmmax', '--maxtempmelting',help='Temperatura de melting máxima.' ,type=int,default=80)
    parser.add_argument('-gcmin', '--minumumgc',help='Porcentaje de GC mínimo.' ,type=int,default=30)
    parser.add_argument('-gcmax', '--maximumgc',help='Porcentaje de GC máximo.' ,type=int,default=70)
    parser.add_argument('-olmin', '--minimumoverlap',help='Sobrelape mínimo.' ,type=int,default=25)
    parser.add_argument('-olmax', '--maximumoverlap',help='Sobrelape máximo.' ,type=int,default=50)
    parser.add_argument('-dghairpin', '--deltaghairpin',help='Mínimo Delta G hairpin permitido.' ,type=int,default=-8000)
    parser.add_argument('-dghomodim', '--deltaghomodimer',help='Mínimo Delta G homodimero permitido.' ,type=int,default=-8000)

    args = parser.parse_args()
    dargs = vars(args)
    rec = descarga.parse_file_to_seqrecord('C:/Users/simon/Downloads/tp53.gb')
    #rec = descarga.accnum_to_seqrecord(dargs["accessionnumber"])
    df = probe_designer(record=rec, transcripts=get_all_transcripts(rec))
                        # ,
                        # minlen=dargs["minimumlength"],
                        # maxlen=dargs["maximumlength"],
                        # tmmin=dargs["mintempmelting"],
                        # tmmax=dargs["maxtempmelting"],
                        # gcmin=dargs["minumumgc"],
                        # gcmax=dargs["maximumgc"],
                        # minoverlap=dargs["minimumoverlap"],
                        # maxoverlap=dargs["maximumoverlap"],
                        # maxoverlap=dargs["deltaghairpin"],
                        # maxoverlap=dargs["deltaghomodimer"])
    
    generate_xlsx(df=df, name='sondas', genes=get_all_genes(rec))
                #   minlen=dargs["minimumlength"],
                #   maxlen=dargs["maximumlength"],
                #   tmmin=dargs["mintempmelting"],
                #   tmmax=dargs["maxtempmelting"],
                #   gcmin=dargs["minumumgc"],
                #   gcmax=dargs["maximumgc"],
                #   minoverlap=dargs["minimumoverlap"],
                #   maxoverlap=dargs["maximumoverlap"])

if __name__ == "__main__":
    main()